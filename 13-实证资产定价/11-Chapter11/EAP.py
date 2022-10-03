import numpy as np
import pandas as pd
import statsmodels.api as sm
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
import openpyxl  # 用于向excel中写入数据

# 相关函数

# 描述性统计
def Summary(data):
    '''
    计算描述性统计

    Parameters
    ----------
    data: 待计算的数据集，要求其中每个df的index为代码，columns为日期  (list)

    Returns
    -------
    描述性统计结果  (pd.DataFrame)
    '''
    result = []
    for i in data:
        des = i.describe().T
        des['Skew'] = i.skew()
        des['Kurt'] = i.kurt()
        des['5%'] = i.quantile(q=0.05)
        des['95%'] = i.quantile(q=0.95)
        result.append(des.mean())
    result = pd.concat(result, axis=1).T
    # 列重排
    result = result[[
        'mean', 'std', 'Skew', 'Kurt', 'min', '5%',
        '25%', '50%', '75%', '95%', 'max', 'count'
    ]]
    # 小数位数
    mapper = {}
    for i in result.columns[:-1]:
        mapper[i] = 2
    mapper['count'] = 0
    result = result.round(mapper)

    return result

# 填充股票代码
def fill_code(data1, data2):
    '''
    以data1的index为基准，对data2填充index
    要求两个数据的index都为股票代码，columns都为日期且相同

    Parameters
    ----------
    data1: 数据1，一般取收益率数据  (pd.DataFrame)
    data2: 数据2，待补充的数据  (pd.DataFrame)

    Returns
    -------
    数据2填充后的结果  (pd.DataFrame)
    '''
    # 为避免data2中有部分股票data1中没有，先对data2取交集
    inter_code = [x for x in data1.index if x in data2.index]
    data2 = data2.loc[inter_code]
    # 补充股票
    temp = pd.DataFrame([], index=data1.index)
    data2 = pd.concat([temp, data2], axis=1, join='outer')

    return data2

# 相关系数
def Corr(data_list1, data_list2, kind):
    '''
    data_list1中每一个数据和data_list2中每一个数据的相关系数
    要求所有数据的index为代码且相同，columns为日期且相同

    Parameters
    ----------
    data_list1: 所研究的变量  (list of pd.DataFrame)
    data_list2: 其他变量  (list of pd.DataFrame)
    kind: 'pearson' or 'spearman'  (str)
    
    Returns
    -------
    相关系数结果  (pd.DataFrame)
    '''
    result = []
    for df1 in data_list1:
        result_i = []
        for df2 in data_list2:
            temp = []
            for m in df1.columns:
                temp.append(df1[m].corr(df2[m], method=kind))  # 每个截面上的相关系数
            result_i.append(np.mean(temp))
        result.append(result_i)
    
    result = pd.DataFrame(result).T

    return result

# 单变量Port分析
def PortChar(X, data_list):
    '''
    X-Sorted Portfolio Characteristics
    要求所有df的index为股票代码且相同，columns为日期且相同

    Parameters
    ----------
    X: 排序变量  (pd.DataFrame)
    data_list: 待计算的特征  (list of pd.DataFrame)

    Returns
    -------
    分组特征计算结果  (pd.DataFrame)
    '''
    result = []
    for df in data_list:
        result_i = []
        for i in X.columns:
            temp = pd.concat([X[i], df[i]], axis=1, keys=['a', 'b'])
            temp = temp[temp['a'].notna()]  # 只保留X为非na的情况
            group = pd.qcut(temp['a'], 10, labels=False)
            result_i.append(temp['b'].groupby(group).mean().tolist())
        result.append(pd.DataFrame(result_i).mean().tolist())
    
    return pd.DataFrame(result)
            
def UniPortR_i(X, Ret, rf, q=10, method='EW', cap=None, all_=True):
    '''
    对于一个异象因子X的Univariate Portfolio Returns
    要求传入的X与ret以及cap，index都为股票代码且相同，
    columns都为日期且相同。其中ret是已经做好ahead计算的收益，因此最后几列为空（取决于ahead几期）

    Parameters
    ----------
    X: 用于排序的变量  (pd.DataFrame)
    Ret: ahead之后的收益  (pd.DataFrame)
    rf: 无风险利率  (pd.Series)
    q: 分为几组  (int, default is 10)
    method: 'EW'代表等权，'VW'代表市值加权  (str)
    cap: 市值  (pd.DataFrame)
    all_: 为True则计算全部组合和多空，否则只计算多空组合  (bool)

    Returns
    -------
    组合收益的结果，同时返回list(顺序为从第一组到第十组)和多空组合收益(index为日期)  (pd.Series)
    all_为False时第一个结果为空list
    '''
    ret = Ret.dropna(axis=1, how='all')  # 删除最后没有收益的几个月
    result = []
    lsret = []  # 多空组合收益
    for i in ret.columns:
        if method == 'EW':  # 若为等权
            temp = pd.concat([X[i], ret[i]], axis=1, keys=['a', 'b']).dropna()
            temp['group'] = pd.qcut(temp['a'], q=q, labels=False)
            if all_:  # 计算十组
                temp1 = (temp.groupby('group')['b'].mean() - rf.shift(-1).loc[i]).tolist()
                lsret.append(temp1[-1] - temp1[0])  # 计算多空
                result.append(temp1)
            else:  # 只计算多空
                lsret.append(
                    temp[temp['group'] == (q-1)]['b'].mean() - temp[temp['group'] == 0]['b'].mean()
                )
        else:  # 若为市值加权
            temp = pd.concat([X[i], ret[i], cap[i]], axis=1, keys=['a', 'b', 'c']).dropna()
            temp['group'] = pd.qcut(temp['a'], q=q, labels=False)
            if all_:
                temp1 = temp.groupby('group').apply(
                    lambda x: (x['b'] * x['c']).sum() / x['c'].sum()
                )
                temp1 = (temp1 - rf.shift(-1).loc[i]).tolist()
                lsret.append(temp1[-1] - temp1[0])
                result.append(temp1)
            else:
                group10 = temp[temp['group'] == (q-1)]
                group1 = temp[temp['group'] == 0]
                lsret.append(((group10['b'] * group10['c']).sum() / group10['c'].sum()) - ((group1['b'] * group1['c']).sum() / group1['c'].sum()))
    if all_:
        result = pd.DataFrame(result).mean().tolist()
    # 将多空组合收益转化为时序
    lsret = pd.Series(lsret, index=X.columns[(len(X.columns)-len(lsret)):])

    return result, lsret

def UniPortR_i_ls(X, Ret, q=10, method='EW', cap=None):
    '''
    对于一个异象因子X的Univariate Portfolio Returns
    要求传入的X与ret以及cap，index都为股票代码且相同，
    columns都为日期且相同。其中ret是已经做好ahead计算的收益，因此最后几列为空（取决于ahead几期）
    只计算多空组合收益

    Parameters
    ----------
    X: 用于排序的变量  (pd.DataFrame)
    Ret: ahead之后的收益  (pd.DataFrame)
    q: 分为几组  (int, default is 10)
    method: 'EW'代表等权，'VW'代表市值加权  (str)
    cap: 市值  (pd.DataFrame)

    Returns
    -------
    多空组合收益  (pd.Series)
    '''
    ret = Ret.dropna(axis=1, how='all')  # 删除最后没有收益的几个月
    lsret = []  # 多空组合收益
    for i in ret.columns:
        if method == 'EW':  # 若为等权
            temp = pd.concat([X[i], ret[i]], axis=1, keys=['a', 'b']).dropna()
            temp['group'] = pd.qcut(temp['a'], q=q, labels=False)
            lsret.append(
                temp[temp['group'] == (q-1)]['b'].mean() - temp[temp['group'] == 0]['b'].mean()
            )
        else:  # 若为市值加权
            temp = pd.concat([X[i], ret[i], cap[i]], axis=1, keys=['a', 'b', 'c']).dropna()
            temp['group'] = pd.qcut(temp['a'], q=q, labels=False)
            group10 = temp[temp['group'] == (q-1)]
            group1 = temp[temp['group'] == 0]
            lsret.append(((group10['b'] * group10['c']).sum() / group10['c'].sum()) - ((group1['b'] * group1['c']).sum() / group1['c'].sum()))
    lsret = pd.Series(lsret, index=X.columns[(len(X.columns)-len(lsret)):])

    return lsret

def NWtest_1sample(a, lags=6):
    '''
    一个序列的NW检验

    Parameters
    ----------
    a: 需要检验的序列  (array-like)
    lags: NW检验的最大滞后阶数  (float)

    Returns
    -------
    序列均值、NW调整后t值  (list)
    '''
    adj_a = np.array(a)
    # 对常数回归
    model = sm.OLS(adj_a, [1] * len(adj_a)).fit(cov_type='HAC', cov_kwds={'maxlags': lags})

    return [adj_a.mean(), float(model.tvalues)]

def RiskAdj(ret, rf, factor_ret, lags=6):
    '''
    计算调整后alpha及对应的NW调整后t值
    要求ret与factor_ret index相同且都为日期，factor_ret是按FF三因子的顺序
    rf的index为日期，与ret格式相同即可

    Parameters
    ----------
    ret: 待调整的收益序列   (pd.Series)
    rf: 无风险利率,只有一列  (pd.Series)
    factor_ret: 三因子收益序列  (pd.DataFrame)
    lags: NW调整的滞后期数  (int, default is 6)

    Returns
    -------
    CAPM调整后alpha、FF调整后alpha、CAPM t、FF t  (list)  
    '''
    excess = ret - rf.loc[ret.index]
    # CAPM调整
    model1 = sm.OLS(excess, sm.add_constant(factor_ret.iloc[:, 0])).fit(
        cov_type='HAC', cov_kwds={'maxlags': lags}
    )
    # FF调整
    model2 = sm.OLS(excess, sm.add_constant(factor_ret)).fit(
        cov_type='HAC', cov_kwds={'maxlags': lags}
    )

    return [model1.params[0], model2.params[0], model1.tvalues[0], model2.tvalues[0]]

def UniPortR(X_list, factor_ret, rf, **kwds):
    '''
    Univariate Portfolio Returns (按书上Table 11.3)
    要求X_list中的所有数据和**kwds中的index都为股票代码且相同，columns都为日期且相同
    factor_ret是所有日期的因子收益，且按FF因子的顺序排列

    Parameters
    ----------
    X_list: 用于分组的异象数据  (list of pd.DataFrame)
    factor_ret: FF因子收益  (pd.DataFrame)
    **kwds: UniPortR_i的参数
    
    Returns
    -------
    结果1：10个排序组合的收益  (pd.DataFrame)
    结果2：多空组合、CAPM alpha、FF alpha及对应t值  (pd.DataFrame)
    '''
    temp0 = [UniPortR_i(x, rf=rf, **kwds) for x in X_list]
    result1 = [x[0] for x in temp0]
    data1 = pd.DataFrame(result1)  # 转化成df
    result2 = [x[1] for x in temp0]
    # 对所有多空进行风险调整
    adj = [RiskAdj(x, rf=rf, factor_ret=factor_ret.loc[x.index]) for x in result2]
    adj = np.array(adj).reshape(len(adj) * 2, 2)
    # 多空组合t检验
    lst = [NWtest_1sample(x) for x in result2]
    lst = np.array(lst).reshape(len(lst) * 2, 1)
    # 生成结果2
    data2 = pd.concat([pd.DataFrame(lst), pd.DataFrame(adj)], axis=1)

    return data1, data2

def UnikAhead(X_list, factor_ret, rf, Ret, ahead=range(1, 13), **kwds):
    '''
    Univariate Portfolio Analysis—k-Month-Ahead Returns
    要求FF因子收益的index为日期且和排序变量columns格式相同

    Parameters
    ----------
    X_list: 用于排序的变量  (list of pd.DataFrame)
    factor_ret: FF因子收益，且按FF顺序  (pd.DataFrame)
    rf: 无风险利率  (pd.Series)
    Ret: 月度收益  (pd.DataFrame)
    ahead: ahead期数包含哪些  (array-like)
    **kwds: UniPortR_i的相关参数

    Returns
    -------
    多空平均收益  (pd.DataFrame)
    多空NW t  (pd.DataFrame)
    多空FF alpha  (pd.DataFrame)
    多空FF t  (pd.DataFrame)
    '''
    temp = []
    for X in X_list:
        temp.append([UniPortR_i_ls(X=X, Ret=Ret.shift(periods=-i, axis=1), **kwds) for i in ahead])
    temp = pd.DataFrame(temp)
    # NW t
    NWt = temp.applymap(NWtest_1sample)
    NWt_mean = NWt.applymap(lambda x: x[0])
    NWt_t = NWt.applymap(lambda x: x[1])
    # FF
    FF = temp.applymap(lambda x: RiskAdj(x, rf=rf, factor_ret=factor_ret.loc[x.index]))
    FF_alpha = FF.applymap(lambda x: x[1])
    FF_t = FF.applymap(lambda x: x[3])

    return NWt_mean, NWt_t, FF_alpha, FF_t

# 双变量非独立分组
def BiDep_all(X1, X2, Ret, rf, f, method='EW', q=(5, 5), cap=None):
    '''
    双变量非独立分组，先按X1分组，后按X2分组，计算所有排序组合和多空组合
    要求所有数据集index为股票代码且相同，columns为日期且相同

    Parameters
    ----------
    X1: 第一个分组的变量  (pd.DataFrame)
    X2: 第二个分组的变量  (pd.DataFrame)
    Ret: 已经ahead后的收益，因此最后一列全为空  (pd.DataFrame)
    rf: 无风险利率  (pd.Series)
    f: 因子收益  (pd.DataFrame)
    method: 'EW'为等权, 'VW'为市值加权  (str)
    q: 分组组数  (tuple, default is (5, 5))
    cap: 市值  (pd.DataFrame)

    Returns
    -------
    按书上表格形式返回结果，如Table 11.5  (pd.DataFrame)
    '''
    ret = Ret.dropna(how='all', axis=1)
    results = []
    if method == 'EW':
        for i in ret.columns:
            temp = pd.concat([X1[i], X2[i], ret[i]], axis=1, keys=['x1', 'x2', 'r'])
            temp['g1'] = pd.qcut(temp['x1'], q=q[0], labels=False)
            x1group = [temp[temp['g1'] == j] for j in range(q[0])]  # 5个x1的组
            result_i = []  # 每个截面的结果
            for x1 in x1group:  # 每组可以计算出6个收益
                x1['g2'] = pd.qcut(x1['x2'], q=q[1], labels=False)  # 根据x2来排序
                temp1 = (x1.groupby('g2')['r'].mean() - rf.shift(-1).loc[i]).tolist()
                temp1.append(temp1[-1] - temp1[0])  # 多空组合
                result_i.append(temp1)
            results.append(pd.DataFrame(result_i).T)
    elif method == 'VW':
        for i in ret.columns:
            temp = pd.concat([X1[i], X2[i], ret[i], cap[i]], axis=1, keys=['x1', 'x2', 'r', 'cap'])
            temp['g1'] = pd.qcut(temp['x1'], q=q[0], labels=False)
            x1group = [temp[temp['g1'] == j] for j in range(q[0])]
            result_i = []
            for x1 in x1group:
                x1['g2'] = pd.qcut(x1['x2'], q=q[1], labels=False)
                temp1 = x1.groupby('g2').apply(lambda x: (x['r'] * x['cap']).sum() / x['cap'].sum())
                temp1 = (temp1 - rf.shift(-1).loc[i]).tolist()
                temp1.append(temp1[-1] - temp1[0])
                result_i.append(temp1)
            results.append(pd.DataFrame(result_i).T)
    # 加入平均组
    for temp2 in results:
        temp2['avg'] = temp2.mean(axis=1)
    # 将25个组合和多空组合分割
    sortp = [x.iloc[:q[1], :] for x in results]
    lsp = [x.iloc[q[1], :].tolist() for x in results]
    lsp = pd.DataFrame(lsp, index=X1.columns[(len(X1.columns)-len(lsp)):])
    # 25个组合的平均收益
    sort_mean = (sum(sortp) / len(ret.columns)).values.tolist()
    # 多空组合NW t和风险调整
    NW = lsp.apply(NWtest_1sample)
    adj = lsp.apply(lambda x: RiskAdj(x, rf=rf, factor_ret=f.loc[x.index]))
    sort_mean.append([x[0] for x in NW])
    sort_mean.append(['(' + str(round(x[1], 2)) + ')' for x in NW])
    sort_mean.append([round(x[0], 2) for x in adj])
    sort_mean.append(['(' + str(round(x[2], 2)) + ')' for x in adj])
    sort_mean.append([round(x[1], 2) for x in adj])
    sort_mean.append(['(' + str(round(x[3], 2)) + ')' for x in adj])

    return pd.DataFrame(sort_mean)

def BiDep_ls(X1, X2, Ret, rf, f, method='EW', q=(5, 5), cap=None):
    '''
    双变量非独立分组，只计算多空组合

    Parameters
    ----------
    X1: 第一个分组的变量  (pd.DataFrame)
    X2: 第二个分组的变量  (pd.DataFrame)
    Ret: 已经ahead后的收益，因此最后一列全为空  (pd.DataFrame)
    rf: 无风险利率  (pd.Series)
    f: 因子收益  (pd.DataFrame)
    method: 'EW'为等权, 'VW'为市值加权  (str)
    q: 分组组数  (tuple, default is (5, 5))
    cap: 市值  (pd.DataFrame)

    Returns
    -------
    平均收益  
    NW t  
    CAPM alpha  
    CAPM t  
    FF alpha  
    FF t  组合df
    '''
    ret = Ret.dropna(how='all', axis=1)
    results = []
    if method == 'EW':
        for i in ret.columns:
            temp = pd.concat([X1[i], X2[i], ret[i]], axis=1, keys=['x1', 'x2', 'r'])
            temp['g1'] = pd.qcut(temp['x1'], q=q[0], labels=False)
            x1group = [temp[temp['g1'] == j] for j in range(q[0])]  # 5个x1的组
            result_i = []  # 每个截面的结果
            for x1 in x1group:  # 每组算出一个多空组合收益
                x1['g2'] = pd.qcut(x1['x2'], q=q[1], labels=False)  # 根据x2来排序
                result_i.append(
                    x1[x1['g2'] == (q[1]-1)]['r'].mean() - x1[x1['g2'] == 0]['r'].mean()
                )
            result_i.append(np.mean(result_i))
            results.append(result_i)
    elif method == 'VW':
        for i in ret.columns:
            temp = pd.concat([X1[i], X2[i], ret[i], cap[i]], axis=1, keys=['x1', 'x2', 'r', 'cap'])
            temp['g1'] = pd.qcut(temp['x1'], q=q[0], labels=False)
            x1group = [temp[temp['g1'] == j] for j in range(q[0])]
            result_i = []
            for x1 in x1group:
                x1['g2'] = pd.qcut(x1['x2'], q=q[1], labels=False)
                grouph = x1[x1['g2'] == (q[1]-1)]
                groupl = x1[x1['g2'] == 0]
                result_i.append(
                    (grouph['r'] * grouph['cap']).sum() / grouph['cap'].sum() - (groupl['r'] * groupl['cap']).sum() / groupl['cap'].sum()
                )
            result_i.append(np.mean(result_i))
            results.append(result_i)
    results = pd.DataFrame(results, index=X1.columns[(len(X1.columns)-len(results)):])
    # NW 
    NW = results.apply(NWtest_1sample)
    # CAPM FF
    adj = results.apply(lambda x: RiskAdj(x, rf=rf, factor_ret=f.loc[x.index]))
    final_result = pd.DataFrame([
        [round(x[0], 2) for x in NW],
        ['(' + str(round(x[1], 2)) + ')' for x in NW],
        [round(x[0], 2) for x in adj],
        ['(' + str(round(x[2], 2)) + ')' for x in adj],
        [round(x[1], 2) for x in adj],
        ['(' + str(round(x[3], 2)) + ')' for x in adj]
    ])

    return final_result

# 双变量独立分组
def BiInd(X1, X2, Ret, rf, f, method='EW', q=(5, 5), cap=None):
    '''
    双变量独立分组，只计算多空组合，按课本表11.7的形式返回结果

    Parameters
    ----------
    X1: 第一个分组的变量，是控制变量  (pd.DataFrame)
    X2: 第二个分组的变量  (pd.DataFrame)
    Ret: 已经ahead后的收益，因此最后一列全为空  (pd.DataFrame)
    rf: 无风险利率  (pd.Series)
    f: 因子收益  (pd.DataFrame)
    method: 'EW'为等权, 'VW'为市值加权  (str)
    q: 分组组数  (tuple, default is (5, 5))
    cap: 市值  (pd.DataFrame)

    Returns
    -------
    平均收益  
    NW t  
    CAPM alpha  
    CAPM t  
    FF alpha  
    FF t  组合df，都保留2位小数
    '''
    ret = Ret.dropna(how='all', axis=1)
    results = []
    if method == 'EW':
        for i in ret.columns:
            temp = pd.concat([X1[i], X2[i], ret[i]], axis=1, keys=['x1', 'x2', 'r'])
            temp['g1'] = pd.qcut(temp['x1'], q=q[0], labels=False)
            temp['g2'] = pd.qcut(temp['x2'], q=q[1], labels=False)
            result_i = []
            for j in range(q[0]):
                temp1 = temp[temp['g1'] == j]
                high = temp1[temp1['g2'] == (q[1]-1)]
                low = temp1[temp1['g2'] == 0]
                if (len(high) > 0) & (len(low) > 0):
                    result_i.append(high['r'].mean() - low['r'].mean())
                else:
                    result_i.append(0)
            result_i.append(np.mean(result_i))
            results.append(result_i)
    elif method == 'VW':
        for i in ret.columns:
            temp = pd.concat([X1[i], X2[i], ret[i], cap[i]], axis=1, keys=['x1', 'x2', 'r', 'cap'])
            temp['g1'] = pd.qcut(temp['x1'], q=q[0], labels=False)
            temp['g2'] = pd.qcut(temp['x2'], q=q[1], labels=False)
            result_i = []
            for j in range(q[0]):
                temp1 = temp[temp['g1'] == j]
                high = temp1[temp1['g2'] == (q[1]-1)]
                low = temp1[temp1['g2'] == 0]
                if (len(high) > 0) & (len(low) > 0):
                    result_i.append(
                        (high['r'] * high['cap']).sum() / high['cap'].sum() - (low['r'] * low['cap']).sum() / low['cap'].sum()
                    )
                else:
                    result_i.append(0)
            result_i.append(np.mean(result_i))
            results.append(result_i)
    results = pd.DataFrame(results, index=X1.columns[(len(X1.columns)-len(results)):])
    # NW 
    NW = results.apply(NWtest_1sample)
    # CAPM FF
    adj = results.apply(lambda x: RiskAdj(x, rf=rf, factor_ret=f.loc[x.index]))
    final_result = pd.DataFrame([
        [round(x[0], 2) for x in NW],
        ['(' + str(round(x[1], 2)) + ')' for x in NW],
        [round(x[0], 2) for x in adj],
        ['(' + str(round(x[2], 2)) + ')' for x in adj],
        [round(x[1], 2) for x in adj],
        ['(' + str(round(x[3], 2)) + ')' for x in adj]
    ])

    return final_result
    



# FM回归
def FMRegression(Ret, rf, X_list, pos, num, level=0.005):
    '''
    给定方程形式的FM回归，在返回的list结果中指定的位置插入空白
    要求收益与所有X的index为股票代码且相同，columns为日期且相同

    Parameters
    ----------
    Ret: 已经ahead之后的月度收益  (pd.DataFrame)
    rf: 无风险利率  (pd.Series)
    X_list: 自变量  (list of df)
    pos: 指定插入位置，要按python的位置  (list of int)
    num: 指定插入空白的个数  (list of int)
    level: 自变量缩尾水平，是绝对量，不是百分比  (float, default is 0.005)

    Returns
    -------
    FM回归的结果，t值保留两位小数且加了括号  (list)
    '''
    # 每个截面回归一次
    ret = Ret.dropna(how='all', axis=1)
    coefs = []  # 系数
    adj = []  # 调整后R2
    n = []  # 样本量
    for i in ret.columns:
        temp = pd.concat([x[i] for x in X_list], axis=1)  # 所有自变量
        temp['r'] = ret[i] - rf.shift(-1).loc[i]
        temp.dropna(inplace=True)
        temp.iloc[:, :-1] = temp.iloc[:, :-1].apply(lambda x: x.clip(
            np.percentile(x, level * 100),
            np.percentile(x, (1 - level) * 100)
        ))  # 缩尾
        model = sm.OLS(temp['r'].values, sm.add_constant(temp.iloc[:, :-1]).values).fit()  # 回归
        coefs.append(model.params)
        adj.append(model.rsquared_adj)
        n.append(len(temp))
    # 系数 NW t
    NW = pd.DataFrame(coefs).apply(NWtest_1sample)
    # 不包含截距项的系数结果
    result = []
    for j in NW.iloc[1:]:
        result.append(round(j[0], 3))
        result.append('(' + str(round(j[1], 2)) + ')')
    # 截距项结果
    result.append(round(NW.iloc[0][0], 3))
    result.append('(' + str(round(NW.iloc[0][1], 2)) + ')')
    result.append(round(np.mean(adj), 2))
    result.append(int(np.mean(n)))
    # 加入空值
    for k in range(len(pos)):
        for m in range(num[k]):
            result.insert(pos[k], np.nan)
    
    return result





# 相关写入excel函数
def ExcelWrite1(data, row, col, file_name, sheet_name):
    '''
    针对只有一个数据集的数据写入

    Parameters
    ----------
    data: 需要写入的数据集  (pd.DataFrame)
    row: 指定开始行  (int)
    col: 指定开始列  (int)
    file_name: excel路径  (str)
    sheet_name: 要写入的sheet的名称  (str)
    '''
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    for i in range(len(data)):
        for j in range(len(data.columns)):
            sheet.cell(row=row+i, column=col+j, value=data.iloc[i, j])
    workbook.save(file_name)

def ExcelWrite2(data1, data2, row, col, file_name, sheet_name):
    '''
    针对同时有结果和t检验的数据写入 (例如表11.3 Panel B)

    Parameters
    ----------
    data1: 10个组合的结果  (pd.DataFrame)
    data2: 风险调整相关结果  (pd.DataFrame)
    row: 指定开始行  (int)
    col: 指定开始列  (int)
    file_name: excel路径  (str)
    sheet_name: 要写入的sheet的名称  (str)
    '''
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    for i in range(len(data1)):  # 此时是写入第一个数据
        for j in range(len(data1.columns)):
            sheet.cell(row=row+2*i, column=col+j, value=data1.iloc[i, j])
    for i in range(len(data2)):  # 此时写入第二个数据
        if i % 2 == 0:  # 非t值则直接写入
            for j in range(len(data2.columns)):
                sheet.cell(row=row+i, column=col+len(data1.columns)+j, value=data2.iloc[i, j])
        else:
            for j in range(len(data2.columns)):
                sheet.cell(row=row+i, column=col+len(data1.columns)+j, value='('+str(data2.iloc[i, j])+')')
    workbook.save(file_name)

def ExcelWrite3(data_list1, data_list2, row, col, file_name, sheet_name):
    '''
    针对一行数值，一行p值的形式写入(例如课本表11.4)
    data_list1中是所有数值的结果，data_list2中每个df对应data_list1中的t值
    且所有df的index相同

    Parameters
    ----------
    data_list1: 所有数值结果  (list of df)
    data_list2: 所有t值结果  (list of df)
    row: 指定开始行  (int)
    col: 指定开始列  (int)
    file_name: excel路径  (str)
    sheet_name: 要写入的sheet的名称  (str)
    '''
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    # 将数值和t值分别合并成一个df，并reshape
    data1 = pd.DataFrame(pd.concat(data_list1, axis=1).values.reshape((len(data_list1[0]) * len(data_list1)), len(data_list1[0].columns)))
    data2 = pd.DataFrame(pd.concat(data_list2, axis=1).values.reshape((len(data_list2[0]) * len(data_list2)), len(data_list2[0].columns)))
    # 将数值和t值合并成一个df，并reshape
    data = pd.DataFrame(pd.concat([data1, data2], axis=1).values.reshape(len(data1) * 2, len(data1.columns)))
    # 写入
    for i in range(len(data)):  # 此时写入第二个数据
        if i % 2 == 0:  # 非t值则直接写入
            for j in range(len(data.columns)):
                sheet.cell(row=row+i, column=col+j, value=data.iloc[i, j])
        else:
            for j in range(len(data.columns)):
                sheet.cell(row=row+i, column=col+j, value='('+str(data.iloc[i, j])+')')
    workbook.save(file_name)

# 画净值图
def plot_Cum_Ret(factor_ret, y, size=(10, 8)):
    '''
    画出累计收益图

    Parameters
    ----------
    factor_ret: 因子月度收益(0.03代表是3%)  (array-like, 最好是np.array)
    y: y轴标签  (str)
    '''
    comp = (factor_ret + 1).cumprod() - 1
    cumlog = np.log((factor_ret + 1).tolist()).cumsum()

    fig = plt.figure(figsize=size)
    ax1 = fig.add_subplot(111)
    lns1=ax1.plot(comp*100, '-k', label = 'Compounded excess return (left axis)')
    plt.xticks(rotation=90)
    ax2 = ax1.twinx()
    lns2= ax2.plot(cumlog*100, '--r', label = 'Cumulative log excess return (right axis)')
    lns = lns1+lns2
    plt.xticks(range(0,241,24),range(2000,2021,2))
    labs = [l.get_label() for l in lns]
    ax1.legend(lns, labs, loc=0)
    fmt='%.0f%%'
    yticks = mtick.FormatStrFormatter(fmt)
    ax1.yaxis.set_major_formatter(yticks)
    ax2.yaxis.set_major_formatter(yticks)
    ax1.set_xlabel('Date')
    ax1.set_ylabel(y)

